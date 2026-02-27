"""
Vibe Document Renderer — lightweight HTTP service for PDF/PNG/XLSX generation.

Endpoints:
  POST /render       — render HTML to PDF or PNG
  POST /xlsx         — generate styled XLSX from columns + rows JSON
  GET  /health       — health check
"""

import io
import unicodedata
import json
import logging
import sys
from flask import Flask, request, jsonify, send_file
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter

app = Flask(__name__)
logging.basicConfig(stream=sys.stderr, level=logging.INFO,
                    format="[DocRenderer] %(message)s")
log = logging.getLogger(__name__)

# ── Modern XLSX Styles ──

# Header style — dark indigo gradient look
# Tahoma has built-in Persian/Arabic glyphs and renders correctly in Excel
HEADER_FONT = Font(name="Tahoma", size=12, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
HEADER_BORDER = Border(
    left=Side(style="thin", color="0F2640"),
    right=Side(style="thin", color="0F2640"),
    top=Side(style="thin", color="0F2640"),
    bottom=Side(style="medium", color="0F2640"),
)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)

# Data rows — clean modern look
DATA_FONT = Font(name="Tahoma", size=11, color="1A1A2E")
DATA_FONT_ALT = Font(name="Tahoma", size=11, color="1A1A2E")
DATA_BORDER = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
DATA_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)
DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)

# Total/summary row style
TOTAL_FONT = Font(name="Tahoma", size=11, bold=True, color="1B3A5C")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_BORDER = Border(
    left=Side(style="thin", color="B0C8DE"),
    right=Side(style="thin", color="B0C8DE"),
    top=Side(style="medium", color="1B3A5C"),
    bottom=Side(style="medium", color="1B3A5C"),
)

# Alternating row fills for readability
EVEN_ROW_FILL = PatternFill(start_color="F8F9FC", end_color="F8F9FC", fill_type="solid")
ODD_ROW_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Title row style (for optional title row above headers)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1B3A5C")
TITLE_ALIGN = Alignment(horizontal="center", vertical="center")

# ── HTML template for PDF/PNG export ──

EXPORT_HTML_TEMPLATE = """<!DOCTYPE html>
<html dir="rtl" lang="fa">
<head>
<meta charset="UTF-8"/>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
@page {{ size: A4; margin: 15mm; }}
body {{
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Tahoma, Arial, sans-serif;
  font-size: 13px;
  line-height: 1.5;
  color: #1a1a1a;
  direction: rtl;
  text-align: right;
}}
h1 {{
  font-size: 20px;
  font-weight: 700;
  color: #1B3A5C;
  margin-bottom: 12px;
  padding-bottom: 8px;
  border-bottom: 2px solid #1B3A5C;
}}
.meta {{
  font-size: 11px;
  color: #666;
  margin-bottom: 14px;
}}
table {{
  width: 100%;
  border-collapse: collapse;
  margin-top: 10px;
}}
th {{
  background: #1B3A5C;
  color: #fff;
  font-weight: 600;
  padding: 8px 10px;
  text-align: right;
  border: 1px solid #0F2640;
  white-space: nowrap;
}}
td {{
  padding: 6px 10px;
  border: 1px solid #d0d5dd;
  text-align: right;
  vertical-align: top;
}}
tr:nth-child(even) td {{ background: #f8f9fc; }}
</style>
</head>
<body>
<h1>{title}</h1>
<div class="meta">{meta}</div>
<table>
<thead><tr>{header_cells}</tr></thead>
<tbody>
{body_rows}
</tbody>
</table>
</body>
</html>"""


def _escape(text):
    return (str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;"))


def _build_export_html(title, columns, rows, meta=""):
    header_cells = "".join(f"<th>{_escape(c)}</th>" for c in columns)
    body_rows = ""
    for row in rows:
        cells = "".join(f"<td>{_escape(v)}</td>" for v in row)
        body_rows += f"<tr>{cells}</tr>\n"
    return EXPORT_HTML_TEMPLATE.format(
        title=_escape(title),
        meta=_escape(meta),
        header_cells=header_cells,
        body_rows=body_rows,
    )


# ── Column width auto-sizing heuristic ──

def _display_width(text):
    """Estimate display width: Arabic/Persian/CJK chars count as ~1.8, Latin as 1."""
    w = 0
    for ch in str(text):
        cat = unicodedata.category(ch)
        if unicodedata.east_asian_width(ch) in ('W', 'F'):
            w += 2.0  # CJK wide chars
        elif '\u0600' <= ch <= '\u06FF' or '\u0750' <= ch <= '\u077F' or '\uFB50' <= ch <= '\uFDFF' or '\uFE70' <= ch <= '\uFEFF':
            w += 1.6  # Arabic/Persian script
        else:
            w += 1.0
    return w


def _auto_column_width(col_name, col_values, min_width=10, max_width=26):
    """Calculate a compact column width based on content."""
    header_w = _display_width(col_name)
    # Sample up to 30 values for width estimation
    sample = col_values[:30]
    if sample:
        # Use the 80th percentile content width (not max) to avoid outlier stretching
        widths = sorted([_display_width(v) for v in sample])
        p80_idx = min(int(len(widths) * 0.8), len(widths) - 1)
        content_w = widths[p80_idx]
    else:
        content_w = 0
    # Use the wider of header or content, with small padding
    width = max(header_w, content_w) + 2
    return max(min_width, min(width, max_width))


# ── Routes ──

@app.route("/health", methods=["GET"])
def health():
    return jsonify(status="ok"), 200


@app.route("/render", methods=["POST"])
def render():
    """Render HTML string or structured data to PDF or PNG."""
    data = request.get_json(force=True)
    fmt = data.get("format", "pdf").lower()
    title = data.get("title", "Document")

    # Accept either raw html or structured columns+rows
    html_content = data.get("html")
    if not html_content:
        columns = data.get("columns", [])
        rows = data.get("rows", [])
        meta = data.get("meta", "")
        if not columns:
            return jsonify(error="Provide 'html' or 'columns'+'rows'"), 400
        html_content = _build_export_html(title, columns, rows, meta)

    try:
        html_obj = HTML(string=html_content)
        buf = io.BytesIO()

        if fmt == "png":
            html_obj.write_png(buf)
            mimetype = "image/png"
            ext = "png"
        else:
            html_obj.write_pdf(buf)
            mimetype = "application/pdf"
            ext = "pdf"

        buf.seek(0)
        log.info(f"Rendered {fmt} ({buf.getbuffer().nbytes} bytes)")
        return send_file(buf, mimetype=mimetype,
                         download_name=f"{title}.{ext}", as_attachment=False)
    except Exception as e:
        log.error(f"Render failed: {e}")
        return jsonify(error=str(e)), 500


@app.route("/xlsx", methods=["POST"])
def xlsx():
    """Generate a professionally styled XLSX file from columns + rows."""
    data = request.get_json(force=True)
    columns = data.get("columns", [])
    rows = data.get("rows", [])
    title = data.get("title", "Spreadsheet")
    sheet_rtl = data.get("rtl", True)

    if not columns:
        return jsonify(error="'columns' is required"), 400

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]
        ws.sheet_view.rightToLeft = sheet_rtl

        # ── RTL column reversal ─────────────────────────────────────────────
        # In RTL mode, Excel shows column A on the RIGHT side of the screen.
        # So for proper RTL layout (first column on the right), we reverse
        # the column order physically: last logical column becomes column A.
        # This matches the HTML/PDF output where the first column is on the right.
        if sheet_rtl:
            display_columns = list(reversed(columns))
            display_rows = [list(reversed(row)) if row else row for row in rows]
        else:
            display_columns = list(columns)
            display_rows = [list(row) if row else row for row in rows]

        num_cols = len(display_columns)

        # Keywords that mark a total/summary row — must match as whole word/phrase,
        # not as a substring (e.g. "کل" must NOT match "سکله")
        TOTAL_KEYWORDS_EXACT = [
            "مجموع", "جمع کل", "مجموع کل", "total", "sum", "subtotal",
            "grand total", "خلاصه"
        ]
        # Keywords that must appear at the START of the cell text
        TOTAL_KEYWORDS_PREFIX = [
            "جمع کل", "مجموع", "جمع:", "total", "sum"
        ]

        def is_total_row(row_data):
            """Detect total/summary rows without false positives."""
            for cell in row_data:
                text = str(cell).strip().lower()
                if not text:
                    continue
                # Exact match: entire cell is a total keyword
                if text in TOTAL_KEYWORDS_EXACT:
                    return True
                # Prefix match: cell starts with a total keyword
                for kw in TOTAL_KEYWORDS_PREFIX:
                    if text.startswith(kw):
                        return True
            return False

        # ── Auto-size column widths ──
        for i, col in enumerate(display_columns, 1):
            letter = get_column_letter(i)
            col_values = [row[i - 1] if i - 1 < len(row) else "" for row in display_rows]
            ws.column_dimensions[letter].width = _auto_column_width(col, col_values)

        # ── Set default row height ──
        ws.sheet_properties.defaultRowHeight = 22

        # ── Header row ──
        header_row_num = 1
        ws.row_dimensions[header_row_num].height = 32

        for col_idx, col_name in enumerate(display_columns, 1):
            cell = ws.cell(row=header_row_num, column=col_idx, value=str(col_name))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = HEADER_BORDER
            cell.alignment = HEADER_ALIGN

        # ── Data rows with alternating colors & total detection ──
        for row_idx, row_data in enumerate(display_rows, 2):
            total_row = is_total_row(row_data)
            is_even = (row_idx % 2 == 0)

            if total_row:
                row_fill = TOTAL_FILL
                row_font = TOTAL_FONT
                row_border = TOTAL_BORDER
                row_height = 28
            else:
                row_fill = EVEN_ROW_FILL if is_even else ODD_ROW_FILL
                row_font = DATA_FONT
                row_border = DATA_BORDER
                row_height = 24

            ws.row_dimensions[row_idx].height = row_height

            for col_idx, value in enumerate(row_data, 1):
                if col_idx > num_cols:
                    break
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.font = row_font
                cell.border = row_border
                cell.alignment = DATA_ALIGN
                cell.fill = row_fill

            # Fill empty trailing cells
            for col_idx in range(len(row_data) + 1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx, value="")
                cell.border = row_border
                cell.alignment = DATA_ALIGN
                cell.fill = row_fill

        # ── Freeze header row for scrolling ──
        ws.freeze_panes = "A2"

        # ── Auto-filter on all columns (only for non-empty data) ──
        if display_rows:
            last_col_letter = get_column_letter(num_cols)
            last_row = len(display_rows) + 1
            ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

        # ── Print settings ──
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        log.info(f"Generated XLSX: {len(display_columns)} cols, {len(display_rows)} rows, {buf.getbuffer().nbytes} bytes")
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            download_name=f"{title}.xlsx",
            as_attachment=False,
        )
    except Exception as e:
        log.error(f"XLSX generation failed: {e}")
        return jsonify(error=str(e)), 500


if __name__ == "__main__":
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 5050
    log.info(f"Starting doc renderer on port {port}")

    # Use waitress for production-grade WSGI server
    try:
        from waitress import serve
        log.info("Using waitress WSGI server")
        serve(app, host="127.0.0.1", port=port, threads=4, channel_timeout=120)
    except ImportError:
        log.warning("waitress not available, falling back to Flask dev server")
        app.run(host="127.0.0.1", port=port, debug=False)
